from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO
from os import cpu_count
from threading import Lock, Thread
from time import perf_counter
from typing import Any
import unicodedata
from uuid import uuid4

from django.core.files.uploadedfile import UploadedFile
from django.db import transaction
from django.db.models import Max, Min, Sum
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from apps.cadastros.models import (
    Cliente,
    FormaPagamento,
    FormaPagamentoMapeamento,
    FormaPagamentoOrigem,
    Produto,
    UnidadeMedida,
    Usuario,
)
from apps.integracao.hash_engine import percentual_semelhanca_textual
from apps.validacao.services_legacy import contar_pendencias_validacao
from apps.validacao.models import STG_AuditoriaPlanilha, STG_ItemVenda, STG_PagamentoVenda, STG_Venda
from apps.vendas.models import ItemVenda, PagamentoVenda, Venda


HOST_VENDA_SHEET_NAME = "HostVenda"
VALID_EXTENSIONS = {".xlsx", ".xlsm"}
DECIMAL_TOLERANCE = Decimal("0.5")
LIMIAR_SEMELHANCA_NOME_PRODUTO = 0.7
LIMIAR_SEMELHANCA_NOME_CLIENTE = 0.7
MOTIVOS_DIVERGENCIA_VALIDOS = {
    "divergencia_totais",
    "divergencia_formato",
    "duplicado_sot",
}
MOTIVOS_DIVERGENCIA_NAO_BLOQUEANTES_CONSOLIDACAO = {
    "divergencia_totais",
    "duplicado_sot",
}
CODIGOS_PRECHECK_COM_OVERRIDE_FORMATO = {
    "divergencia_formato_pagamento",
}
MOTIVOS_RECONCILIACAO_COM_OVERRIDE_FORMATO = {
    "divergencia_formato",
}
UNIDADE_ALIAS: dict[str, str] = {
    "UND": "UN",
    "UNID": "UN",
    "UNIDADE": "UN",
    "LITRO": "LT",
    "LITROS": "LT",
    "LTS": "LT",
    "KILO": "KG",
    "KILOGRAMA": "KG",
    "KILOGRAMAS": "KG",
    "GRAMA": "G",
    "GRAMAS": "G",
}

_IMPORT_JOBS: dict[str, dict[str, Any]] = {}
_IMPORT_JOBS_LOCK = Lock()


@dataclass
class ImportErrorItem:
    arquivo: str
    linha: int
    motivo: str


@dataclass
class ValidationResult:
    vendas_aprovadas: int
    vendas_divergentes: int
    vendas_duplicadas_sot: int
    divergencias: list[dict[str, Any]]
    inconsistencias: list[dict[str, Any]]
    kpis: dict[str, Any]
    pode_aprovar: bool


class ReconciliacaoBloqueioError(ValueError):
    def __init__(
        self,
        detail: str,
        *,
        codigo: str,
        bloqueios: list[dict[str, Any]] | None = None,
        permite_override: bool = False,
    ) -> None:
        super().__init__(detail)
        self.detail = detail
        self.codigo = codigo
        self.bloqueios = bloqueios or []
        self.permite_override = bool(permite_override)

    def to_payload(self) -> dict[str, Any]:
        return {
            "detail": self.detail,
            "codigo": self.codigo,
            "bloqueios": self.bloqueios,
            "permite_override": self.permite_override,
        }


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_identity_text(value: Any) -> str:
    raw = _normalize_text(value)
    if not raw:
        return ""
    normalized = unicodedata.normalize("NFKD", raw)
    without_accents = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return " ".join(without_accents.upper().split())


def _normalize_unit_token(value: Any) -> str:
    text = _normalize_identity_text(value)
    return "".join(ch for ch in text if ch.isalnum())


def _resolver_unidade_legado(
    *,
    unidade_legado: str,
    unidades_por_sigla: dict[str, UnidadeMedida],
) -> UnidadeMedida | None:
    token = _normalize_unit_token(unidade_legado)
    if not token:
        return None

    unidade = unidades_por_sigla.get(token)
    if unidade is not None:
        return unidade

    alias = UNIDADE_ALIAS.get(token)
    if alias:
        return unidades_por_sigla.get(alias)
    return None


def _formatar_erros_precheck(erros: list[str], max_items: int = 40) -> str:
    linhas = ["Consolidacao bloqueada por inconsistencias de pre-check."]
    for item in erros[:max_items]:
        linhas.append(f"- {item}")

    restante = len(erros) - max_items
    if restante > 0:
        linhas.append(f"- ... e mais {restante} inconsistencia(s).")

    return "\n".join(linhas)


def _nomes_clientes_semelhantes(nome_stg: Any, nome_sot: Any) -> tuple[bool, float]:
    nome_stg_norm = _normalize_identity_text(nome_stg)
    nome_sot_norm = _normalize_identity_text(nome_sot)
    if not nome_stg_norm or not nome_sot_norm:
        return True, 1.0

    similaridade = percentual_semelhanca_textual(nome_stg, nome_sot)
    return similaridade >= LIMIAR_SEMELHANCA_NOME_CLIENTE, similaridade


def _nomes_produtos_semelhantes(nome_stg: Any, nome_sot: Any) -> tuple[bool, float]:
    nome_stg_norm = _normalize_identity_text(nome_stg)
    nome_sot_norm = _normalize_identity_text(nome_sot)
    if not nome_stg_norm or not nome_sot_norm:
        return True, 1.0

    similaridade = percentual_semelhanca_textual(
        nome_stg,
        nome_sot,
        ordenar_tokens=True,
    )
    return similaridade >= LIMIAR_SEMELHANCA_NOME_PRODUTO, similaridade


def _motivos_snapshot_venda(venda: STG_Venda) -> set[str]:
    snapshot = venda.snapshot_divergencia or {}
    motivos_raw = snapshot.get("motivos") or []
    return {
        _normalize_text(item).lower()
        for item in motivos_raw
        if _normalize_text(item)
    }


def _motivos_bloqueantes_consolidacao(venda: STG_Venda) -> set[str]:
    return _motivos_snapshot_venda(venda) - MOTIVOS_DIVERGENCIA_NAO_BLOQUEANTES_CONSOLIDACAO


def _formatar_erros_validacao_bloqueada(erros: list[str], max_items: int = 20) -> str:
    linhas = ["Validacao bloqueada por inconsistencias estruturais."]
    for item in erros[:max_items]:
        linhas.append(f"- {item}")

    restante = len(erros) - max_items
    if restante > 0:
        linhas.append(f"- ... e mais {restante} inconsistencia(s).")

    return "\n".join(linhas)


def _extrair_codigo_precheck(erro: str) -> str:
    erro_norm = _normalize_text(erro)
    if not erro_norm:
        return "inconsistencia"

    _, separador, resto = erro_norm.partition(":")
    trecho = resto if separador else erro_norm
    token = _normalize_text(trecho).split(" ", 1)[0].strip().lower()
    return token or "inconsistencia"


def _codigos_precheck_apenas_formato(codigos: set[str]) -> bool:
    return bool(codigos) and codigos <= CODIGOS_PRECHECK_COM_OVERRIDE_FORMATO


def _motivos_reconciliacao_apenas_formato(motivos: set[str]) -> bool:
    return bool(motivos) and motivos <= MOTIVOS_RECONCILIACAO_COM_OVERRIDE_FORMATO


def _permite_override_precheck(codigos: set[str], forcar_divergencia_formato: bool) -> bool:
    return bool(forcar_divergencia_formato) and _codigos_precheck_apenas_formato(codigos)


def _permite_override_motivos_reconciliacao(
    motivos: set[str],
    forcar_divergencia_formato: bool,
) -> bool:
    return bool(forcar_divergencia_formato) and _motivos_reconciliacao_apenas_formato(motivos)


def _montar_bloqueio_precheck_venda(venda: STG_Venda, erros_precheck: list[str]) -> dict[str, Any]:
    codigos = {
        _extrair_codigo_precheck(item)
        for item in erros_precheck
        if _normalize_text(item)
    }
    codigos_ordenados = sorted(codigos)

    return {
        "tipo_documento": venda.tipo_documento,
        "id_legado": int(venda.id_legado),
        "venda": f"{venda.tipo_documento} #{int(venda.id_legado):06d}",
        "erros": erros_precheck[:20],
        "codigos": codigos_ordenados,
        "possui_divergencia_formato": "divergencia_formato_pagamento" in codigos,
        "permite_override": _codigos_precheck_apenas_formato(codigos),
    }


def _agrupar_erros_precheck_por_venda(erros: list[str]) -> list[dict[str, Any]]:
    agrupado: dict[tuple[str, int], dict[str, Any]] = {}
    gerais: list[str] = []

    for erro in erros:
        erro_norm = _normalize_text(erro)
        if not erro_norm:
            continue

        prefixo_raw, separador, _ = erro_norm.partition(":")
        if not separador or "#" not in prefixo_raw:
            gerais.append(erro_norm)
            continue

        tipo_raw, id_legado_raw = prefixo_raw.split("#", 1)
        tipo_documento = _normalize_text(tipo_raw).upper()
        try:
            id_legado = int(_normalize_text(id_legado_raw))
        except ValueError:
            gerais.append(erro_norm)
            continue

        chave = (tipo_documento, id_legado)
        registro = agrupado.setdefault(
            chave,
            {
                "tipo_documento": tipo_documento,
                "id_legado": id_legado,
                "venda": f"{tipo_documento} #{id_legado:06d}",
                "erros": [],
                "_codigos": set(),
            },
        )
        if len(registro["erros"]) < 20:
            registro["erros"].append(erro_norm)
        registro["_codigos"].add(_extrair_codigo_precheck(erro_norm))

    bloqueios: list[dict[str, Any]] = []
    for registro in agrupado.values():
        codigos = sorted(registro.pop("_codigos"))
        bloqueios.append(
            {
                **registro,
                "codigos": codigos,
                "possui_divergencia_formato": "divergencia_formato_pagamento" in codigos,
                "permite_override": _codigos_precheck_apenas_formato(set(codigos)),
            }
        )

    if gerais:
        bloqueios.append(
            {
                "tipo_documento": "",
                "id_legado": None,
                "venda": "GERAL",
                "erros": gerais[:20],
                "codigos": ["inconsistencia_geral"],
                "possui_divergencia_formato": False,
                "permite_override": False,
            }
        )

    bloqueios.sort(key=lambda item: (str(item.get("tipo_documento") or ""), int(item.get("id_legado") or 0)))
    return bloqueios[:200]


def _coletar_erros_precheck_consolidacao(candidatas: list[STG_Venda]) -> list[str]:
    if not candidatas:
        return []

    itens_por_venda: dict[int, list[STG_ItemVenda]] = {}
    pagamentos_por_venda: dict[int, list[STG_PagamentoVenda]] = {}
    produto_ids: set[int] = set()
    usuario_ids: set[int] = set()
    cliente_ids: set[int] = set()
    tipos_documento: set[str] = set()
    ids_legado: set[int] = set()

    for stg_venda in candidatas:
        itens = list(stg_venda.itens.all())
        pagamentos = list(stg_venda.pagamentos.all())
        itens_por_venda[stg_venda.pk] = itens
        pagamentos_por_venda[stg_venda.pk] = pagamentos
        tipos_documento.add(stg_venda.tipo_documento)
        ids_legado.add(int(stg_venda.id_legado))

        if stg_venda.id_usuario_legado is not None:
            usuario_ids.add(int(stg_venda.id_usuario_legado))

        if stg_venda.id_cliente_legado not in (None, 0):
            cliente_ids.add(int(stg_venda.id_cliente_legado))

        for item in itens:
            if item.id_produto_legado is not None:
                produto_ids.add(int(item.id_produto_legado))

    usuarios_por_id = {
        int(usuario.id_usuario): usuario
        for usuario in Usuario.objects.filter(id_usuario__in=usuario_ids)
    }
    clientes_por_id = {
        int(cliente.id_cliente): cliente
        for cliente in Cliente.objects.filter(id_cliente__in=cliente_ids)
    }
    produtos_por_id = {
        int(produto.id_produto): produto
        for produto in Produto.objects.filter(id_produto__in=produto_ids)
    }

    unidades_por_sigla: dict[str, UnidadeMedida] = {}
    for unidade in UnidadeMedida.objects.all():
        token = _normalize_unit_token(unidade.sigla)
        if token:
            unidades_por_sigla[token] = unidade

    formas_origem_validas = {
        (str(tipo).strip().upper(), int(id_origem))
        for tipo, id_origem in FormaPagamentoOrigem.objects.filter(ativo=True).values_list(
            "tipo_documento", "id_forma_origem"
        )
    }
    mapeamentos_ativos = {
        (str(item.tipo_documento).strip().upper(), int(item.id_forma_origem)): item.forma_pagamento
        for item in FormaPagamentoMapeamento.objects.select_related("forma_pagamento").filter(ativo=True)
    }

    auditoria_pagamentos_por_venda: dict[tuple[str, int], set[str]] = {}
    for row in STG_AuditoriaPlanilha.objects.filter(
        tipo_documento__in=tipos_documento,
        id_legado__in=ids_legado,
    ).values("tipo_documento", "id_legado", "tipo_pagamento_descricao"):
        key = (row["tipo_documento"], int(row["id_legado"]))
        agg = auditoria_pagamentos_por_venda.setdefault(key, set())
        agg.update(_extract_pagamento_tokens(row["tipo_pagamento_descricao"]))

    clientes_padrao = list(Cliente.objects.filter(cliente_padrao=True).order_by("id_cliente")[:2])
    if len(clientes_padrao) > 1:
        return ["Consolidacao bloqueada: existe mais de um cliente padrao configurado."]
    cliente_padrao = clientes_padrao[0] if clientes_padrao else None

    erros_precheck: list[str] = []
    for stg_venda in candidatas:
        prefixo = f"{stg_venda.tipo_documento}#{stg_venda.id_legado}"
        chave_venda = (stg_venda.tipo_documento, int(stg_venda.id_legado))

        itens = itens_por_venda.get(stg_venda.pk, [])
        pagamentos = pagamentos_por_venda.get(stg_venda.pk, [])
        if not itens or not pagamentos:
            erros_precheck.append(f"{prefixo}: venda_sem_itens_ou_pagamentos")
            continue

        pagamentos_stg = {
            token
            for token in {
                _normalize_compare_token(pg.tipo_pagamento_descricao_legado)
                for pg in pagamentos
            }
            if token
        }
        pagamentos_auditoria = auditoria_pagamentos_por_venda.get(chave_venda, set())
        if pagamentos_stg != pagamentos_auditoria:
            erros_precheck.append(
                f"{prefixo}: divergencia_formato_pagamento "
                f"(stg='{sorted(pagamentos_stg)}', auditoria='{sorted(pagamentos_auditoria)}')"
            )

        if stg_venda.id_usuario_legado is None:
            erros_precheck.append(f"{prefixo}: usuario_legado_ausente")
        else:
            usuario = usuarios_por_id.get(int(stg_venda.id_usuario_legado))
            if usuario is None:
                erros_precheck.append(
                    f"{prefixo}: usuario_legado_nao_encontrado (id={stg_venda.id_usuario_legado})"
                )
            else:
                nome_stg = _normalize_identity_text(stg_venda.nome_usuario_legado)
                nome_sot = _normalize_identity_text(usuario.nome)
                if nome_stg and nome_sot and nome_stg != nome_sot:
                    erros_precheck.append(
                        f"{prefixo}: usuario_nome_divergente "
                        f"(stg='{stg_venda.nome_usuario_legado}', sot='{usuario.nome}')"
                    )

        if stg_venda.id_cliente_legado in (None, 0):
            if cliente_padrao is None:
                erros_precheck.append(f"{prefixo}: cliente_legado_zero_sem_cliente_padrao_configurado")
        else:
            cliente = clientes_por_id.get(int(stg_venda.id_cliente_legado))
            if cliente is None:
                erros_precheck.append(f"{prefixo}: cliente_nao_encontrado (id={stg_venda.id_cliente_legado})")
            else:
                nomes_compativeis, similaridade_cliente = _nomes_clientes_semelhantes(
                    stg_venda.nome_cliente_legado,
                    cliente.nome_cliente,
                )
                if not nomes_compativeis:
                    erros_precheck.append(
                        f"{prefixo}: cliente_nome_divergente "
                        f"(similaridade={similaridade_cliente * 100:.1f}%, "
                        f"stg='{stg_venda.nome_cliente_legado}', sot='{cliente.nome_cliente}')"
                    )

        for item in itens:
            if item.id_produto_legado is None:
                erros_precheck.append(f"{prefixo}: item_sem_id_produto")
                continue

            produto = produtos_por_id.get(int(item.id_produto_legado))
            if produto is None:
                erros_precheck.append(f"{prefixo}: produto_nao_encontrado (id={item.id_produto_legado})")
                continue

            nomes_compativeis, similaridade_produto = _nomes_produtos_semelhantes(
                item.nome_produto_legado,
                produto.produto,
            )
            if not nomes_compativeis:
                erros_precheck.append(
                    f"{prefixo}: produto_nome_divergente "
                    f"(id={item.id_produto_legado}, similaridade={similaridade_produto * 100:.1f}%, "
                    f"stg='{item.nome_produto_legado}', sot='{produto.produto}')"
                )
                continue

            unidade = _resolver_unidade_legado(
                unidade_legado=item.unidade_comercial_legado,
                unidades_por_sigla=unidades_por_sigla,
            )
            if unidade is None:
                erros_precheck.append(
                    f"{prefixo}: unidade_legado_sem_mapeamento "
                    f"(valor='{item.unidade_comercial_legado}')"
                )

        for pg in pagamentos:
            if pg.id_tipo_pagamento_legado is None:
                erros_precheck.append(f"{prefixo}: pagamento_sem_id_forma_origem")
                continue

            chave = (stg_venda.tipo_documento, int(pg.id_tipo_pagamento_legado))
            if chave not in formas_origem_validas:
                erros_precheck.append(
                    f"{prefixo}: forma_origem_nao_cadastrada "
                    f"(tipo={stg_venda.tipo_documento}, id_origem={pg.id_tipo_pagamento_legado})"
                )
                continue

            forma = mapeamentos_ativos.get(chave)
            if forma is None:
                erros_precheck.append(
                    f"{prefixo}: mapeamento_forma_ausente "
                    f"(tipo={stg_venda.tipo_documento}, id_origem={pg.id_tipo_pagamento_legado})"
                )

    return erros_precheck



def _normalize_tipo_documento(value: Any) -> str:
    raw = _normalize_text(value).upper().replace("-", "")
    if raw in {"NFCE", "NFC-E"}:
        return STG_AuditoriaPlanilha.TIPO_NFCE
    if raw == STG_AuditoriaPlanilha.TIPO_DAV:
        return STG_AuditoriaPlanilha.TIPO_DAV
    return ""


def _resolver_forma_por_origem(
    *,
    tipo_documento: str,
    id_forma_origem: int | None,
    descricao_origem: str = "",
    persist_fallback: bool = False,
) -> FormaPagamento | None:
    if id_forma_origem is None:
        return None

    tipo_norm = _normalize_text(tipo_documento).upper()

    mapeamento = (
        FormaPagamentoMapeamento.objects.select_related("forma_pagamento")
        .filter(tipo_documento=tipo_norm, id_forma_origem=id_forma_origem, ativo=True)
        .first()
    )
    if mapeamento is not None:
        return mapeamento.forma_pagamento

    forma = FormaPagamento.objects.filter(id_forma=id_forma_origem).first()
    if forma is None and descricao_origem:
        forma = FormaPagamento.objects.filter(descricao__iexact=_normalize_text(descricao_origem)).first()

    if forma is None:
        return None

    if persist_fallback and tipo_norm:
        FormaPagamentoMapeamento.objects.get_or_create(
            tipo_documento=tipo_norm,
            id_forma_origem=id_forma_origem,
            defaults={
                "forma_pagamento": forma,
                "descricao_origem": _normalize_text(descricao_origem) or _normalize_text(forma.descricao),
                "ativo": True,
            },
        )

    return forma


def _resolver_origem_por_forma(*, tipo_documento: str, id_forma_destino: int) -> tuple[int, str]:
    tipo_norm = _normalize_text(tipo_documento).upper()
    forma = FormaPagamento.objects.filter(id_forma=id_forma_destino).first()
    if forma is None:
        raise ValueError("Forma de pagamento destino nao encontrada.")

    mapeamento = FormaPagamentoMapeamento.objects.filter(
        forma_pagamento_id=id_forma_destino,
        tipo_documento=tipo_norm,
        ativo=True,
    ).first()
    if mapeamento is None:
        raise ValueError(
            f"Nao existe mapeamento para a forma '{forma.descricao}' no tipo '{tipo_norm}'."
        )

    descricao = _normalize_text(mapeamento.descricao_origem) or _normalize_text(forma.descricao)
    return int(mapeamento.id_forma_origem), descricao


def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    raw = _normalize_text(value)
    if not raw:
        return None

    # Aceita entradas com simbolos monetarios e separadores pt-BR/en-US.
    cleaned = raw.replace("R$", "").replace(" ", "")
    allowed = "".join(ch for ch in cleaned if ch.isdigit() or ch in {".", ",", "-"})
    if not allowed or allowed in {"-", ".", ",", "-.", "-,"}:
        return None

    dot_count = allowed.count(".")
    comma_count = allowed.count(",")

    if dot_count and comma_count:
        # O ultimo separador visto e tratado como separador decimal.
        if allowed.rfind(",") > allowed.rfind("."):
            normalized = allowed.replace(".", "").replace(",", ".")
        else:
            normalized = allowed.replace(",", "")
    elif comma_count:
        normalized = allowed.replace(",", ".") if comma_count == 1 else allowed.replace(",", "")
    elif dot_count:
        normalized = allowed if dot_count == 1 else allowed.replace(".", "")
    else:
        normalized = allowed

    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None


def _to_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        except Exception:
            return None

    raw = _normalize_text(value)
    if not raw:
        return None

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _to_time(value: Any) -> time | None:
    if value is None:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.time()
            if isinstance(converted, time):
                return converted
        except Exception:
            return None

    raw = _normalize_text(value)
    if not raw:
        return None

    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    return None


def _normalize_compare_token(value: Any) -> str:
    return _normalize_text(value).upper()


def _extract_pagamento_tokens(value: Any) -> set[str]:
    raw = _normalize_compare_token(value)
    if not raw:
        return set()

    tokens = {raw}
    for sep in ("/", "|", ",", ";", "+"):
        if sep in raw:
            parts = {_normalize_compare_token(part) for part in raw.split(sep)}
            tokens |= {part for part in parts if part}
    return {token for token in tokens if token}


def _find_last_data_row_by_col_d(sheet: Any, min_row: int = 6) -> int:
    last_data_row = 0
    saw_data = False
    empty_streak = 0

    # Emula a ideia do End(xlUp) do VBA: procura a última linha útil da coluna D.
    for row_idx, (id_legado_raw,) in enumerate(
        sheet.iter_rows(min_row=min_row, min_col=4, max_col=4, values_only=True),
        start=min_row,
    ):
        if id_legado_raw not in (None, ""):
            last_data_row = row_idx
            saw_data = True
            empty_streak = 0
            continue

        if saw_data:
            empty_streak += 1
            # Se já vimos dados e em seguida há um bloco grande vazio,
            # interrompe a busca para evitar varrer planilhas enormes vazias.
            if empty_streak >= 5000:
                break

    return last_data_row


def _collect_file_rows_from_payload(filename: str, payload: bytes) -> tuple[list[STG_AuditoriaPlanilha], list[ImportErrorItem]]:
    errors: list[ImportErrorItem] = []
    records: list[STG_AuditoriaPlanilha] = []

    extension = filename.lower().rsplit(".", 1)
    if len(extension) < 2 or f".{extension[-1]}" not in VALID_EXTENSIONS:
        return records, [ImportErrorItem(arquivo=filename, linha=0, motivo="Extensao nao suportada.")]

    try:
        workbook = load_workbook(filename=BytesIO(payload), data_only=True, read_only=True)
    except Exception as exc:
        return records, [ImportErrorItem(arquivo=filename, linha=0, motivo=f"Falha ao abrir arquivo: {exc}")]

    if HOST_VENDA_SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        return records, [ImportErrorItem(arquivo=filename, linha=0, motivo="Aba HostVenda nao encontrada.")]

    sheet = workbook[HOST_VENDA_SHEET_NAME]
    last_row = _find_last_data_row_by_col_d(sheet, min_row=6)

    if last_row < 6:
        workbook.close()
        return records, errors

    for row_offset, row_data in enumerate(
        sheet.iter_rows(min_row=6, max_row=last_row, min_col=4, max_col=25, values_only=True),
        start=6,
    ):
        id_legado_raw = row_data[0]
        if id_legado_raw in (None, ""):
            continue

        data_venda = _to_date(row_data[1])
        hora_venda = _to_time(row_data[3])
        tipo_documento = _normalize_tipo_documento(row_data[4])
        usuario_nome = _normalize_text(row_data[9])
        cliente_nome = _normalize_text(row_data[14])
        tipo_pagamento_descricao = _normalize_text(row_data[18])
        valor_total = _to_decimal(row_data[21])

        try:
            id_legado = int(id_legado_raw)
        except (ValueError, TypeError):
            errors.append(ImportErrorItem(arquivo=filename, linha=row_offset, motivo="ID legado invalido."))
            continue

        if data_venda is None:
            errors.append(ImportErrorItem(arquivo=filename, linha=row_offset, motivo="Data de venda invalida."))
            continue

        if not tipo_documento:
            errors.append(ImportErrorItem(arquivo=filename, linha=row_offset, motivo="Tipo de documento invalido."))
            continue

        if valor_total is None:
            errors.append(ImportErrorItem(arquivo=filename, linha=row_offset, motivo="Valor total invalido."))
            continue

        records.append(
            STG_AuditoriaPlanilha(
                data_venda=data_venda,
                hora_venda=hora_venda,
                tipo_documento=tipo_documento,
                id_legado=id_legado,
                usuario_nome=usuario_nome,
                cliente_nome=cliente_nome,
                valor_total=valor_total,
                tipo_pagamento_descricao=tipo_pagamento_descricao,
            )
        )

    workbook.close()
    return records, errors


def _set_job(job_id: str, **kwargs: Any) -> None:
    with _IMPORT_JOBS_LOCK:
        job = _IMPORT_JOBS.get(job_id)
        if job is None:
            return
        job.update(kwargs)


def get_importacao_job(job_id: str) -> dict[str, Any] | None:
    with _IMPORT_JOBS_LOCK:
        job = _IMPORT_JOBS.get(job_id)
        if job is None:
            return None
        return dict(job)


def _build_divergencia_snapshot(
    *,
    venda: STG_Venda,
    motivos: list[str],
    total_documento: Decimal,
    total_itens: Decimal,
    total_itens_via_fallback: bool,
    total_pagamentos: Decimal,
    status_venda: str,
    auditoria_encontrada: bool,
    auditoria_valor: Decimal,
    auditoria_pagamentos: list[str],
    pagamentos_stg: list[str],
) -> dict[str, Any]:
    return {
        "motivos": motivos,
        "status_venda": status_venda,
        "total_documento": str(total_documento),
        "total_itens": str(total_itens),
        "total_itens_via_fallback": bool(total_itens_via_fallback),
        "total_pagamentos": str(total_pagamentos),
        "total_auditoria": str(auditoria_valor) if auditoria_encontrada else None,
        "formato_venda": pagamentos_stg,
        "formato_auditoria": auditoria_pagamentos,
    }


def _avaliar_venda(
    *,
    venda: STG_Venda,
    total_itens: Decimal,
    total_itens_via_fallback: bool,
    total_pagamentos: Decimal,
    pagamentos_tokens: set[str],
    auditoria_encontrada: bool,
    auditoria_total: Decimal,
    auditoria_pagamentos: set[str],
    sot_keys: set[tuple[str, int]],
) -> tuple[list[str], dict[str, Any]]:
    total_documento = venda.valor_final or Decimal("0")
    motivos: list[str] = []
    status_venda_norm = _normalize_text(venda.status_venda).upper()

    if venda.validacao_override:
        return motivos, {}

    divergencia_totais = (
        abs(total_documento - total_itens) > DECIMAL_TOLERANCE
        or abs(total_documento - total_pagamentos) > DECIMAL_TOLERANCE
        or (not auditoria_encontrada)
        or abs(total_documento - auditoria_total) > DECIMAL_TOLERANCE
        or status_venda_norm == "C"
    )

    if divergencia_totais:
        motivos.append("divergencia_totais")

    if pagamentos_tokens != auditoria_pagamentos:
        motivos.append("divergencia_formato")

    if (venda.tipo_documento, int(venda.id_legado)) in sot_keys:
        motivos.append("duplicado_sot")

    motivos = sorted(set(motivos))
    snapshot = _build_divergencia_snapshot(
        venda=venda,
        motivos=motivos,
        total_documento=total_documento,
        total_itens=total_itens,
        total_itens_via_fallback=total_itens_via_fallback,
        total_pagamentos=total_pagamentos,
        status_venda=venda.status_venda,
        auditoria_encontrada=auditoria_encontrada,
        auditoria_valor=auditoria_total,
        auditoria_pagamentos=sorted(auditoria_pagamentos),
        pagamentos_stg=sorted(pagamentos_tokens),
    )
    return motivos, snapshot


def executar_tripla_validacao(reset_tracking: bool = False) -> ValidationResult:
    if reset_tracking:
        STG_Venda.objects.update(
            validacao_override=False,
            status_tratamento=STG_Venda.TRATAMENTO_PENDENTE,
            snapshot_divergencia={},
            tratamento_atualizado_em=None,
        )

    # Prioriza itens ativos (cancelado=False). Quando uma venda possui somente itens marcados
    # como cancelados, usa fallback para o total de todos os itens para evitar total_itens zerado.
    item_totais_ativos = {
        (item["tipo_documento"], int(item["id_venda_legado"])): item["total_itens"] or Decimal("0")
        for item in STG_ItemVenda.objects.filter(cancelado=False)
        .values("tipo_documento", "id_venda_legado")
        .annotate(total_itens=Sum("valor_total_calculado"))
    }

    item_totais_todos = {
        (item["tipo_documento"], int(item["id_venda_legado"])): item["total_itens"] or Decimal("0")
        for item in STG_ItemVenda.objects.values("tipo_documento", "id_venda_legado").annotate(total_itens=Sum("valor_total_calculado"))
    }

    pagamentos_totais = {
        (item["tipo_documento"], int(item["id_venda_legado"])): item["total_pago"] or Decimal("0")
        for item in STG_PagamentoVenda.objects.values("tipo_documento", "id_venda_legado").annotate(total_pago=Sum("valor_pago"))
    }

    pagamentos_map: dict[tuple[str, int], set[str]] = {}
    for pagamento in STG_PagamentoVenda.objects.values(
        "tipo_documento", "id_venda_legado", "tipo_pagamento_descricao_legado"
    ):
        key = (pagamento["tipo_documento"], int(pagamento["id_venda_legado"]))
        pagamentos_map.setdefault(key, set()).add(
            _normalize_compare_token(pagamento["tipo_pagamento_descricao_legado"])
        )

    auditoria_map: dict[tuple[str, int], dict[str, Any]] = {}
    for row in STG_AuditoriaPlanilha.objects.all():
        key = (row.tipo_documento, int(row.id_legado))
        agg = auditoria_map.setdefault(
            key,
            {
                "encontrada": False,
                "total": Decimal("0"),
                "pagamentos": set(),
                "pagamentos_detalhe": [],
            },
        )
        agg["encontrada"] = True
        total_valor = row.valor_total or Decimal("0")
        agg["total"] += total_valor
        agg["pagamentos"] |= _extract_pagamento_tokens(row.tipo_pagamento_descricao)
        agg["pagamentos_detalhe"].append(
            {
                "tipo_pagamento_descricao": _normalize_text(row.tipo_pagamento_descricao),
                "valor_linha": str(total_valor),
            }
        )

    sot_keys = set(Venda.objects.values_list("tipo_documento", "id_legado"))

    aprovadas = 0
    divergentes = 0
    duplicadas = 0
    detalhes_divergencia: list[dict[str, Any]] = []
    inconsistencias: list[dict[str, Any]] = []

    motivos_count: dict[str, int] = {
        "divergencia_totais": 0,
        "divergencia_formato": 0,
        "duplicado_sot": 0,
    }

    soma_stg_finalizadas = Decimal("0")
    soma_stg_canceladas = Decimal("0")
    soma_auditoria_finalizadas = Decimal("0")
    soma_auditoria_canceladas = Decimal("0")
    soma_validadas_finalizadas = Decimal("0")
    qtd_validadas_finalizadas = 0
    qtd_auditoria_finalizadas = 0

    vendas = list(STG_Venda.objects.all())
    for venda in vendas:
        key = (venda.tipo_documento, int(venda.id_legado))
        total_itens = item_totais_ativos.get(key)
        total_itens_via_fallback = total_itens is None
        if total_itens_via_fallback:
            total_itens = item_totais_todos.get(key, Decimal("0"))
        total_pagamentos = pagamentos_totais.get(key, Decimal("0"))
        total_documento = venda.valor_final or Decimal("0")
        status_venda_norm = _normalize_text(venda.status_venda).upper()
        if status_venda_norm == "C":
            soma_stg_canceladas += total_documento
        else:
            soma_stg_finalizadas += total_documento

        auditoria_data = auditoria_map.get(key) or {
            "encontrada": False,
            "total": Decimal("0"),
            "pagamentos": set(),
        }
        pagamentos = pagamentos_map.get(key, set())
        motivos, snapshot = _avaliar_venda(
            venda=venda,
            total_itens=total_itens,
            total_itens_via_fallback=total_itens_via_fallback,
            total_pagamentos=total_pagamentos,
            pagamentos_tokens=pagamentos,
            auditoria_encontrada=bool(auditoria_data["encontrada"]),
            auditoria_total=auditoria_data["total"],
            auditoria_pagamentos=set(auditoria_data["pagamentos"]),
            sot_keys=sot_keys,
        )

        if auditoria_data["encontrada"]:
            if status_venda_norm == "C":
                soma_auditoria_canceladas += auditoria_data["total"]
            else:
                soma_auditoria_finalizadas += auditoria_data["total"]
                qtd_auditoria_finalizadas += 1

        if "duplicado_sot" in motivos:
            duplicadas += 1

        if motivos:
            if "duplicado_sot" in motivos:
                venda.status_tratamento = STG_Venda.TRATAMENTO_NEGLIGENCIADO
            if venda.status_tratamento != STG_Venda.TRATAMENTO_NEGLIGENCIADO:
                divergentes += 1
            venda.status_validacao = STG_Venda.STATUS_DIVERGENTE
            venda.snapshot_divergencia = snapshot
            clean_motivos = sorted(set(motivos))
            for motivo in clean_motivos:
                motivos_count[motivo] += 1

            inconsistencias.append(
                {
                    "tipo_documento": venda.tipo_documento,
                    "id_legado": venda.id_legado,
                    **snapshot,
                    "tratamento": venda.status_tratamento,
                }
            )
            detalhes_divergencia.append(
                {
                    "tipo_documento": venda.tipo_documento,
                    "id_legado": venda.id_legado,
                    "motivos": clean_motivos,
                }
            )
        else:
            aprovadas += 1
            venda.status_validacao = STG_Venda.STATUS_APROVADO
            venda.snapshot_divergencia = {}
            if not venda.validacao_override:
                venda.status_tratamento = STG_Venda.TRATAMENTO_PENDENTE

        if venda.status_validacao == STG_Venda.STATUS_APROVADO and status_venda_norm != "C":
            qtd_validadas_finalizadas += 1
            soma_validadas_finalizadas += total_documento

    if vendas:
        STG_Venda.objects.bulk_update(vendas, ["status_validacao", "status_tratamento", "snapshot_divergencia"], batch_size=2000)

    kpis = {
        "total_vendas_stg": len(vendas),
        "vendas_aprovadas": aprovadas,
        "vendas_divergentes": divergentes,
        "vendas_duplicadas_sot": duplicadas,
        "motivos_divergencia": motivos_count,
        "soma_valor_stg": str(soma_stg_finalizadas),
        "soma_valor_stg_canceladas": str(soma_stg_canceladas),
        "soma_valor_auditoria": str(soma_auditoria_finalizadas),
        "soma_valor_auditoria_canceladas": str(soma_auditoria_canceladas),
        "qtd_vendas_validadas": qtd_validadas_finalizadas,
        "soma_valor_vendas_validadas": str(soma_validadas_finalizadas),
        "qtd_vendas_auditoria": qtd_auditoria_finalizadas,
        "diferenca_total": str(soma_validadas_finalizadas - soma_auditoria_finalizadas),
        "diferenca_total_stg_auditoria": str(soma_stg_finalizadas - soma_auditoria_finalizadas),
        "vendas_tratadas": STG_Venda.objects.filter(status_validacao=STG_Venda.STATUS_DIVERGENTE)
        .exclude(status_tratamento=STG_Venda.TRATAMENTO_PENDENTE)
        .count(),
        "vendas_negligenciadas": STG_Venda.objects.filter(status_tratamento=STG_Venda.TRATAMENTO_NEGLIGENCIADO).count(),
    }

    periodo = STG_Venda.objects.aggregate(data_inicial=Min("data_venda"), data_final=Max("data_venda"))
    kpis["periodo_data_inicial"] = periodo["data_inicial"].isoformat() if periodo["data_inicial"] else None
    kpis["periodo_data_final"] = periodo["data_final"].isoformat() if periodo["data_final"] else None

    return ValidationResult(
        vendas_aprovadas=aprovadas,
        vendas_divergentes=divergentes,
        vendas_duplicadas_sot=duplicadas,
        divergencias=detalhes_divergencia[:300],
        inconsistencias=inconsistencias[:500],
        kpis=kpis,
        pode_aprovar=divergentes == 0 and aprovadas > 0,
    )


def obter_kpis_reconciliacao() -> dict[str, Any]:
    validation = executar_tripla_validacao(reset_tracking=False)
    return validation.kpis


def listar_divergencias_reconciliacao(
    *,
    motivo: str | None = None,
    status_tratamento: str | None = None,
    somente_finalizados: bool = False,
    status_venda: str | None = None,
    id_legado: str | int | None = None,
) -> list[dict[str, Any]]:
    motivo_norm = _normalize_text(motivo).lower()
    if motivo_norm and motivo_norm not in MOTIVOS_DIVERGENCIA_VALIDOS:
        motivo_norm = ""

    status_norm = _normalize_text(status_tratamento).upper()
    if status_norm not in {
        "",
        STG_Venda.TRATAMENTO_PENDENTE,
        STG_Venda.TRATAMENTO_AJUSTADO,
        STG_Venda.TRATAMENTO_VALIDADO,
        STG_Venda.TRATAMENTO_NEGLIGENCIADO,
    }:
        status_norm = ""

    status_venda_norm = _normalize_text(status_venda).upper()
    if status_venda_norm not in {"", "F", "C"}:
        status_venda_norm = ""

    id_legado_norm = _normalize_text(id_legado)

    divergentes = (
        STG_Venda.objects.filter(status_validacao=STG_Venda.STATUS_DIVERGENTE)
        .prefetch_related("itens", "pagamentos")
        .order_by("tipo_documento", "id_legado")
    )

    if id_legado_norm:
        try:
            divergentes = divergentes.filter(id_legado=int(id_legado_norm))
        except (ValueError, TypeError):
            divergentes = divergentes.none()

    auditoria_rows = STG_AuditoriaPlanilha.objects.values(
        "tipo_documento",
        "id_legado",
        "tipo_pagamento_descricao",
        "valor_total",
    )
    auditoria_details_map: dict[tuple[str, int], list[dict[str, Any]]] = {}
    for row in auditoria_rows:
        key = (row["tipo_documento"], int(row["id_legado"]))
        auditoria_details_map.setdefault(key, []).append(
            {
                "tipo_pagamento_descricao": _normalize_text(row["tipo_pagamento_descricao"]),
                "valor_linha": str(row["valor_total"] or Decimal("0")),
            }
        )

    cache_forma_canonica: dict[tuple[str, int], int | None] = {}
    rows: list[dict[str, Any]] = []
    for venda in divergentes:
        if status_norm and venda.status_tratamento != status_norm:
            continue

        if somente_finalizados and _normalize_text(venda.status_venda).upper() != "F":
            continue

        if status_venda_norm and _normalize_text(venda.status_venda).upper() != status_venda_norm:
            continue

        snapshot = venda.snapshot_divergencia or {}
        motivos = snapshot.get("motivos") or []
        if motivo_norm and motivo_norm not in motivos:
            continue

        pagamentos_detalhe = []
        for pg in venda.pagamentos.all():
            id_forma_canonica = None
            if pg.id_tipo_pagamento_legado is not None:
                chave_cache = (venda.tipo_documento, int(pg.id_tipo_pagamento_legado))
                if chave_cache in cache_forma_canonica:
                    id_forma_canonica = cache_forma_canonica[chave_cache]
                else:
                    forma_canonica = _resolver_forma_por_origem(
                        tipo_documento=venda.tipo_documento,
                        id_forma_origem=int(pg.id_tipo_pagamento_legado),
                        descricao_origem=_normalize_text(pg.tipo_pagamento_descricao_legado),
                    )
                    id_forma_canonica = int(forma_canonica.id_forma) if forma_canonica is not None else None
                    cache_forma_canonica[chave_cache] = id_forma_canonica

            pagamentos_detalhe.append(
                {
                    "id_stg_pagamento_venda": pg.id_stg_pagamento_venda,
                    "id_tipo_pagamento_legado": pg.id_tipo_pagamento_legado,
                    "id_forma_canonica": id_forma_canonica,
                    "tipo_pagamento_descricao_legado": pg.tipo_pagamento_descricao_legado,
                    "valor_pago": str(pg.valor_pago or Decimal("0")),
                }
            )

        rows.append(
            {
                "tipo_documento": venda.tipo_documento,
                "id_legado": venda.id_legado,
                "data_venda": venda.data_venda.isoformat() if venda.data_venda else None,
                "venda": f"{venda.tipo_documento} #{int(venda.id_legado):06d}",
                "motivos": motivos,
                "tratamento": venda.status_tratamento,
                "status_venda": venda.status_venda,
                "nome_cliente_legado": venda.nome_cliente_legado,
                "stg": {
                    "valor_documento": snapshot.get("total_documento"),
                    "total_itens": snapshot.get("total_itens"),
                    "total_pagamentos": snapshot.get("total_pagamentos"),
                    "pagamentos": snapshot.get("formato_venda") or [],
                    "itens": [
                        {
                            "id_stg_item_venda": item.id_stg_item_venda,
                            "status_venda": item.status_venda,
                            "id_produto_legado": item.id_produto_legado,
                            "nome_produto_legado": item.nome_produto_legado,
                            "quantidade": str(item.quantidade or Decimal("0")),
                            "valor_unitario": str(item.valor_unitario or Decimal("0")),
                            "valor_total_calculado": str(item.valor_total_calculado or Decimal("0")),
                            "cancelado": item.cancelado,
                        }
                        for item in venda.itens.all()
                    ],
                    "pagamentos_detalhe": pagamentos_detalhe,
                },
                "auditoria": {
                    "valor_total": snapshot.get("total_auditoria"),
                    "pagamentos": snapshot.get("formato_auditoria") or [],
                    "pagamentos_detalhe": auditoria_details_map.get((venda.tipo_documento, int(venda.id_legado)), []),
                },
                "totais": {
                    "total_documento": snapshot.get("total_documento"),
                    "total_itens": snapshot.get("total_itens"),
                    "total_itens_via_fallback": bool(snapshot.get("total_itens_via_fallback")),
                    "total_pagamentos": snapshot.get("total_pagamentos"),
                    "total_auditoria": snapshot.get("total_auditoria"),
                },
                "atualizado_em": venda.tratamento_atualizado_em.isoformat() if venda.tratamento_atualizado_em else None,
            }
        )

    return rows


def aplicar_tratamento_divergencia(
    *,
    tipo_documento: str,
    id_legado: int,
    acao: str,
    payload: dict[str, Any],
) -> dict[str, Any]:
    venda = STG_Venda.objects.filter(tipo_documento=tipo_documento, id_legado=id_legado).first()
    if venda is None:
        raise ValueError("Venda STG nao encontrada para tratamento.")

    acao_norm = _normalize_text(acao).lower()
    if acao_norm not in {"validar", "ajustar", "negligenciar"}:
        raise ValueError("Acao invalida. Use 'validar', 'ajustar' ou 'negligenciar'.")

    forcar_divergencia_formato = bool(payload.get("forcar_divergencia_formato"))

    with transaction.atomic():
        if acao_norm == "validar":
            erros_precheck = _coletar_erros_precheck_consolidacao([venda])
            if erros_precheck:
                bloqueio = _montar_bloqueio_precheck_venda(venda, erros_precheck)
                codigos = set(bloqueio.get("codigos") or [])
                if not _permite_override_precheck(codigos, forcar_divergencia_formato):
                    raise ReconciliacaoBloqueioError(
                        _formatar_erros_validacao_bloqueada(erros_precheck),
                        codigo="validacao_bloqueada_precheck",
                        bloqueios=[bloqueio],
                        permite_override=bool(bloqueio.get("permite_override")),
                    )

            venda.validacao_override = True
            venda.status_validacao = STG_Venda.STATUS_APROVADO
            venda.status_tratamento = STG_Venda.TRATAMENTO_VALIDADO
            venda.snapshot_divergencia = {}
            venda.tratamento_atualizado_em = timezone.now()
            venda.save(
                update_fields=[
                    "validacao_override",
                    "status_validacao",
                    "status_tratamento",
                    "snapshot_divergencia",
                    "tratamento_atualizado_em",
                ]
            )
        elif acao_norm == "negligenciar":
            venda.validacao_override = False
            venda.status_validacao = STG_Venda.STATUS_DIVERGENTE
            venda.status_tratamento = STG_Venda.TRATAMENTO_NEGLIGENCIADO
            venda.tratamento_atualizado_em = timezone.now()
            venda.save(
                update_fields=[
                    "validacao_override",
                    "status_validacao",
                    "status_tratamento",
                    "tratamento_atualizado_em",
                ]
            )
        else:
            venda.validacao_override = False
            update_fields_venda = ["validacao_override", "tratamento_atualizado_em"]

            valor_final_raw = payload.get("valor_final")
            if valor_final_raw not in (None, ""):
                valor_final = _to_decimal(valor_final_raw)
                if valor_final is None:
                    raise ValueError("valor_final invalido.")
                venda.valor_final = valor_final
                update_fields_venda.append("valor_final")

            status_venda_raw = _normalize_text(payload.get("status_venda")).upper()
            if status_venda_raw:
                if status_venda_raw not in {"F", "C"}:
                    raise ValueError("status_venda invalido. Use 'F' ou 'C'.")
                venda.status_venda = status_venda_raw
                update_fields_venda.append("status_venda")
                STG_ItemVenda.objects.filter(stg_venda=venda).update(status_venda=status_venda_raw)

            for item_data in payload.get("itens") or []:
                item_id = item_data.get("id_stg_item_venda")
                item = STG_ItemVenda.objects.filter(id_stg_item_venda=item_id, stg_venda=venda).first()
                if item is None:
                    continue

                for field_name in ("quantidade", "valor_unitario", "valor_total_calculado"):
                    if field_name in item_data:
                        value = _to_decimal(item_data.get(field_name))
                        setattr(item, field_name, value)

                if "cancelado" in item_data:
                    item.cancelado = bool(item_data.get("cancelado"))

                item.save(update_fields=["quantidade", "valor_unitario", "valor_total_calculado", "cancelado"])

            for pg_data in payload.get("pagamentos") or []:
                pg_id = pg_data.get("id_stg_pagamento_venda")
                pagamento = STG_PagamentoVenda.objects.filter(id_stg_pagamento_venda=pg_id, stg_venda=venda).first()
                if pagamento is None:
                    continue

                if "valor_pago" in pg_data:
                    valor_pago = _to_decimal(pg_data.get("valor_pago"))
                    if valor_pago is not None:
                        pagamento.valor_pago = valor_pago

                if "id_tipo_pagamento_legado" in pg_data:
                    pagamento.id_tipo_pagamento_legado = pg_data.get("id_tipo_pagamento_legado")

                id_forma_canonica_raw = pg_data.get("id_forma_canonica")
                if id_forma_canonica_raw not in (None, ""):
                    try:
                        id_forma_canonica = int(id_forma_canonica_raw)
                    except (TypeError, ValueError):
                        raise ValueError("id_forma_canonica invalido.")

                    id_origem_resolvido, descricao_origem = _resolver_origem_por_forma(
                        tipo_documento=venda.tipo_documento,
                        id_forma_destino=id_forma_canonica,
                    )
                    pagamento.id_tipo_pagamento_legado = id_origem_resolvido
                    pagamento.tipo_pagamento_descricao_legado = descricao_origem

                if "tipo_pagamento_descricao_legado" in pg_data:
                    pagamento.tipo_pagamento_descricao_legado = _normalize_text(pg_data.get("tipo_pagamento_descricao_legado"))

                pagamento.save(
                    update_fields=["valor_pago", "id_tipo_pagamento_legado", "tipo_pagamento_descricao_legado"]
                )

            venda.tratamento_atualizado_em = timezone.now()
            venda.save(update_fields=update_fields_venda)

    validation = executar_tripla_validacao(reset_tracking=False)
    return {
        "detail": "Tratamento aplicado com sucesso.",
        "kpis": validation.kpis,
    }


def aplicar_tratamento_divergencias_lote(
    *,
    vendas: list[dict[str, Any]],
    acao: str,
    payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    acao_norm = _normalize_text(acao).lower()
    if acao_norm not in {"validar", "negligenciar", "editar_pagamento"}:
        raise ValueError("Acao em lote invalida. Use 'validar', 'negligenciar' ou 'editar_pagamento'.")

    payload = payload or {}
    forcar_divergencia_formato = bool(payload.get("forcar_divergencia_formato"))
    id_forma_destino: int | None = None
    if acao_norm == "editar_pagamento":
        id_forma_raw = payload.get("id_forma")
        if id_forma_raw in (None, ""):
            raise ValueError("Para editar em lote, informe id_forma no payload.")
        try:
            id_forma_destino = int(id_forma_raw)
        except (TypeError, ValueError):
            raise ValueError("id_forma invalido para edicao em lote.")

        forma = FormaPagamento.objects.filter(id_forma=id_forma_destino).first()
        if forma is None:
            raise ValueError("Forma de pagamento destino nao encontrada.")

    processadas = 0
    ignoradas = 0
    bloqueadas = 0
    bloqueios: list[dict[str, Any]] = []
    with transaction.atomic():
        for venda_data in vendas:
            tipo_documento = _normalize_text(venda_data.get("tipo_documento")).upper()
            id_legado_raw = venda_data.get("id_legado")
            try:
                id_legado = int(id_legado_raw)
            except (TypeError, ValueError):
                ignoradas += 1
                continue

            venda = STG_Venda.objects.filter(tipo_documento=tipo_documento, id_legado=id_legado).first()
            if venda is None:
                ignoradas += 1
                continue

            if acao_norm == "validar":
                erros_precheck = _coletar_erros_precheck_consolidacao([venda])
                if erros_precheck:
                    bloqueio = _montar_bloqueio_precheck_venda(venda, erros_precheck)
                    codigos = set(bloqueio.get("codigos") or [])
                    if not _permite_override_precheck(codigos, forcar_divergencia_formato):
                        bloqueadas += 1
                        bloqueios.append(bloqueio)
                        continue

                venda.validacao_override = True
                venda.status_validacao = STG_Venda.STATUS_APROVADO
                venda.status_tratamento = STG_Venda.TRATAMENTO_VALIDADO
                venda.snapshot_divergencia = {}
                venda.tratamento_atualizado_em = timezone.now()
                venda.save(
                    update_fields=[
                        "validacao_override",
                        "status_validacao",
                        "status_tratamento",
                        "snapshot_divergencia",
                        "tratamento_atualizado_em",
                    ]
                )
            elif acao_norm == "negligenciar":
                venda.validacao_override = False
                venda.status_validacao = STG_Venda.STATUS_DIVERGENTE
                venda.status_tratamento = STG_Venda.TRATAMENTO_NEGLIGENCIADO
                venda.tratamento_atualizado_em = timezone.now()
                venda.save(
                    update_fields=[
                        "validacao_override",
                        "status_validacao",
                        "status_tratamento",
                        "tratamento_atualizado_em",
                    ]
                )
            else:
                id_origem_resolvido, descricao_origem = _resolver_origem_por_forma(
                    tipo_documento=venda.tipo_documento,
                    id_forma_destino=int(id_forma_destino),
                )
                STG_PagamentoVenda.objects.filter(stg_venda=venda).update(
                    id_tipo_pagamento_legado=id_origem_resolvido,
                    tipo_pagamento_descricao_legado=descricao_origem,
                )

            processadas += 1

    validation = executar_tripla_validacao(reset_tracking=False)
    return {
        "detail": "Tratamento em lote concluido.",
        "processadas": processadas,
        "ignoradas": ignoradas,
        "bloqueadas": bloqueadas,
        "bloqueios": bloqueios[:200],
        "kpis": validation.kpis,
    }


def listar_formas_pagamento() -> list[dict[str, Any]]:
    formas = FormaPagamento.objects.order_by("descricao", "id_forma")
    return [{"id_forma": int(item.id_forma), "descricao": item.descricao} for item in formas]


def _importar_planilhas_auditoria_sync(payloads: list[dict[str, Any]], job_id: str | None = None) -> dict[str, Any]:
    started_at = perf_counter()
    imported_records: list[STG_AuditoriaPlanilha] = []
    import_errors: list[ImportErrorItem] = []
    erros_precheck_inicial: list[str] = []

    total_files = len(payloads)
    parse_started_at = perf_counter()
    parse_workers = max(1, min(total_files, cpu_count() or 4, 8))

    with ThreadPoolExecutor(max_workers=parse_workers) as executor:
        future_to_file: dict[Any, str] = {}
        for file_info in payloads:
            filename = _normalize_text(file_info.get("name") or "arquivo_sem_nome")
            content = file_info.get("content") or b""
            future = executor.submit(_collect_file_rows_from_payload, filename, content)
            future_to_file[future] = filename

        processed = 0
        for future in as_completed(future_to_file):
            filename = future_to_file[future]
            try:
                records, errors = future.result()
            except Exception as exc:
                records = []
                errors = [
                    ImportErrorItem(
                        arquivo=filename,
                        linha=0,
                        motivo=f"Falha no parser paralelo: {exc}",
                    )
                ]

            imported_records.extend(records)
            import_errors.extend(errors)
            processed += 1

            if job_id:
                _set_job(
                    job_id,
                    stage="importando_arquivos",
                    detail=f"Processando {processed}/{total_files}: {filename}",
                    arquivos_processados=processed,
                    total_arquivos=total_files,
                    linhas_parciais=len(imported_records),
                )

    parse_elapsed_s = perf_counter() - parse_started_at

    db_started_at = perf_counter()
    with transaction.atomic():
        STG_AuditoriaPlanilha.objects.all().delete()

        if imported_records:
            STG_AuditoriaPlanilha.objects.bulk_create(imported_records, batch_size=3000)

        if job_id:
            _set_job(
                job_id,
                stage="validando",
                detail="Executando tripla validacao entre STG e auditoria.",
            )

        validation = executar_tripla_validacao(reset_tracking=True)

        candidatas_precheck = list(
            STG_Venda.objects.filter(status_validacao=STG_Venda.STATUS_APROVADO)
            .exclude(status_tratamento=STG_Venda.TRATAMENTO_NEGLIGENCIADO)
            .prefetch_related("itens", "pagamentos")
            .order_by("tipo_documento", "id_legado")
        )
        existentes_sot = set(Venda.objects.values_list("tipo_documento", "id_legado"))
        candidatas_precheck = [
            venda
            for venda in candidatas_precheck
            if (venda.tipo_documento, int(venda.id_legado)) not in existentes_sot
        ]

        erros_precheck_inicial = _coletar_erros_precheck_consolidacao(candidatas_precheck)

    db_and_validation_elapsed_s = perf_counter() - db_started_at
    total_elapsed_s = perf_counter() - started_at

    return {
        "arquivos_recebidos": total_files,
        "linhas_importadas": len(imported_records),
        "performance": {
            "parse_workers": parse_workers,
            "parse_segundos": round(parse_elapsed_s, 3),
            "persistencia_validacao_segundos": round(db_and_validation_elapsed_s, 3),
            "total_segundos": round(total_elapsed_s, 3),
        },
        "erros_importacao": [
            {"arquivo": err.arquivo, "linha": err.linha, "motivo": err.motivo}
            for err in import_errors[:500]
        ],
        "validacao": {
            "vendas_aprovadas": validation.vendas_aprovadas,
            "vendas_divergentes": validation.vendas_divergentes,
            "vendas_duplicadas_sot": validation.vendas_duplicadas_sot,
            "divergencias": validation.divergencias,
            "inconsistencias": validation.inconsistencias,
            "kpis": validation.kpis,
            "pode_aprovar": validation.pode_aprovar,
        },
        "precheck_inicial": {
            "total_inconsistencias": len(erros_precheck_inicial),
            "inconsistencias": erros_precheck_inicial[:500],
            "bloqueante_nesta_etapa": False,
        },
    }


def _process_import_job(job_id: str, payloads: list[dict[str, Any]]) -> None:
    try:
        _set_job(job_id, status="processing", stage="iniciando", detail="Iniciando importacao de planilhas.")
        resultado = _importar_planilhas_auditoria_sync(payloads, job_id=job_id)
        _set_job(
            job_id,
            status="completed",
            stage="concluido",
            detail="Importacao e validacao concluidas.",
            resultado=resultado,
        )
    except Exception as exc:
        _set_job(
            job_id,
            status="failed",
            stage="erro",
            detail=f"Falha no processamento: {exc}",
            erro=str(exc),
        )


def start_importacao_planilhas_auditoria(uploads: list[UploadedFile]) -> str:
    if not uploads:
        raise ValueError("Nenhum arquivo foi enviado para importacao.")

    payloads: list[dict[str, Any]] = []
    for upload in uploads:
        payloads.append(
            {
                "name": _normalize_text(getattr(upload, "name", "arquivo_sem_nome")),
                "content": upload.read(),
            }
        )

    job_id = uuid4().hex
    with _IMPORT_JOBS_LOCK:
        _IMPORT_JOBS[job_id] = {
            "job_id": job_id,
            "status": "queued",
            "stage": "fila",
            "detail": "Aguardando processamento.",
            "arquivos_processados": 0,
            "total_arquivos": len(payloads),
            "linhas_parciais": 0,
            "resultado": None,
            "erro": None,
        }

    worker = Thread(target=_process_import_job, args=(job_id, payloads), daemon=True)
    worker.start()
    return job_id


def limpar_fluxo_reconciliacao() -> dict[str, int]:
    with transaction.atomic():
        stg_pagamentos_removidos = STG_PagamentoVenda.objects.count()
        stg_itens_removidos = STG_ItemVenda.objects.count()
        stg_vendas_removidas = STG_Venda.objects.count()
        stg_auditoria_removidas = STG_AuditoriaPlanilha.objects.count()

        STG_PagamentoVenda.objects.all().delete()
        STG_ItemVenda.objects.all().delete()
        STG_Venda.objects.all().delete()
        STG_AuditoriaPlanilha.objects.all().delete()

    return {
        "stg_vendas_removidas": stg_vendas_removidas,
        "stg_itens_removidos": stg_itens_removidos,
        "stg_pagamentos_removidos": stg_pagamentos_removidos,
        "stg_auditoria_removidas": stg_auditoria_removidas,
    }


def consolidar_stg_para_sot(*, forcar_divergencia_formato: bool = False) -> dict[str, Any]:
    divergentes_pendentes = list(
        STG_Venda.objects.filter(status_validacao=STG_Venda.STATUS_DIVERGENTE)
        .exclude(status_tratamento=STG_Venda.TRATAMENTO_NEGLIGENCIADO)
        .prefetch_related("itens", "pagamentos")
        .order_by("tipo_documento", "id_legado")
    )
    divergencias_bloqueantes = []
    for venda in divergentes_pendentes:
        motivos_bloqueantes = sorted(_motivos_bloqueantes_consolidacao(venda))
        if motivos_bloqueantes:
            divergencias_bloqueantes.append(
                f"{venda.tipo_documento}#{venda.id_legado}: {', '.join(motivos_bloqueantes)}"
            )

    if divergencias_bloqueantes:
        bloqueios_reconciliacao: list[dict[str, Any]] = []
        motivos_bloqueantes_globais: set[str] = set()
        for venda in divergentes_pendentes:
            motivos_bloqueantes = sorted(_motivos_bloqueantes_consolidacao(venda))
            if not motivos_bloqueantes:
                continue
            motivos_bloqueantes_globais.update(motivos_bloqueantes)
            bloqueios_reconciliacao.append(
                {
                    "tipo_documento": venda.tipo_documento,
                    "id_legado": int(venda.id_legado),
                    "venda": f"{venda.tipo_documento} #{int(venda.id_legado):06d}",
                    "erros": [f"{venda.tipo_documento}#{venda.id_legado}: {', '.join(motivos_bloqueantes)}"],
                    "codigos": motivos_bloqueantes,
                    "possui_divergencia_formato": "divergencia_formato" in motivos_bloqueantes,
                    "permite_override": _motivos_reconciliacao_apenas_formato(set(motivos_bloqueantes)),
                }
            )

        if not _permite_override_motivos_reconciliacao(
            motivos_bloqueantes_globais,
            forcar_divergencia_formato,
        ):
            raise ReconciliacaoBloqueioError(
                "Consolidacao bloqueada por divergencias estruturais na reconciliacao:\n"
                + "\n".join(f"- {item}" for item in divergencias_bloqueantes[:40]),
                codigo="consolidacao_bloqueada_divergencia_reconciliacao",
                bloqueios=bloqueios_reconciliacao[:200],
                permite_override=_motivos_reconciliacao_apenas_formato(motivos_bloqueantes_globais),
            )

    pendencias = contar_pendencias_validacao()
    if any(pendencias.values()):
        raise ValueError(
            "Consolidacao bloqueada por pendencias de cadastro: "
            f"produtos={pendencias['produtos']}, "
            f"clientes={pendencias['clientes']}, "
            f"fornecedores={pendencias['fornecedores']}."
        )

    aprovadas = list(
        STG_Venda.objects.filter(status_validacao=STG_Venda.STATUS_APROVADO).exclude(
            status_tratamento=STG_Venda.TRATAMENTO_NEGLIGENCIADO
        )
        .prefetch_related("itens", "pagamentos")
        .order_by("tipo_documento", "id_legado")
    )

    candidatas_divergentes: list[STG_Venda] = []
    for venda in divergentes_pendentes:
        motivos_bloqueantes = _motivos_bloqueantes_consolidacao(venda)
        if not motivos_bloqueantes:
            candidatas_divergentes.append(venda)
            continue

        if _permite_override_motivos_reconciliacao(motivos_bloqueantes, forcar_divergencia_formato):
            candidatas_divergentes.append(venda)

    if candidatas_divergentes:
        aprovadas.extend(candidatas_divergentes)

    vendas_para_consolidar_por_chave: dict[tuple[str, int], STG_Venda] = {}
    for venda in aprovadas:
        chave = (venda.tipo_documento, int(venda.id_legado))
        vendas_para_consolidar_por_chave[chave] = venda

    aprovadas = list(vendas_para_consolidar_por_chave.values())

    if not aprovadas:
        raise ValueError("Nao ha vendas aprovadas na STG para consolidar.")

    existentes = set(Venda.objects.values_list("tipo_documento", "id_legado"))

    ignoradas_duplicadas = 0
    candidatas: list[STG_Venda] = []
    itens_por_venda: dict[int, list[STG_ItemVenda]] = {}
    pagamentos_por_venda: dict[int, list[STG_PagamentoVenda]] = {}

    produto_ids: set[int] = set()
    usuario_ids: set[int] = set()
    cliente_ids: set[int] = set()

    for stg_venda in aprovadas:
        key = (stg_venda.tipo_documento, int(stg_venda.id_legado))
        if key in existentes:
            ignoradas_duplicadas += 1
            continue

        candidatas.append(stg_venda)
        itens = list(stg_venda.itens.all())
        pagamentos = list(stg_venda.pagamentos.all())
        itens_por_venda[stg_venda.pk] = itens
        pagamentos_por_venda[stg_venda.pk] = pagamentos

        if stg_venda.id_usuario_legado is not None:
            usuario_ids.add(int(stg_venda.id_usuario_legado))

        if stg_venda.id_cliente_legado not in (None, 0):
            cliente_ids.add(int(stg_venda.id_cliente_legado))

        for item in itens:
            if item.id_produto_legado is not None:
                produto_ids.add(int(item.id_produto_legado))

    if not candidatas:
        with transaction.atomic():
            stg_vendas_removidas = STG_Venda.objects.count()
            stg_auditoria_removidas = STG_AuditoriaPlanilha.objects.count()
            STG_Venda.objects.all().delete()
            STG_AuditoriaPlanilha.objects.all().delete()

        return {
            "vendas_inseridas": 0,
            "vendas_ignoradas_duplicadas": ignoradas_duplicadas,
            "vendas_ignoradas_incompletas": 0,
            "detalhes_ignorados": [],
            "momento_consolidacao": None,
            "stg_vendas_removidas": stg_vendas_removidas,
            "stg_auditoria_removidas": stg_auditoria_removidas,
        }

    usuarios_por_id = {
        int(usuario.id_usuario): usuario
        for usuario in Usuario.objects.filter(id_usuario__in=usuario_ids)
    }
    clientes_por_id = {
        int(cliente.id_cliente): cliente
        for cliente in Cliente.objects.filter(id_cliente__in=cliente_ids)
    }
    produtos_por_id = {
        int(produto.id_produto): produto
        for produto in Produto.objects.filter(id_produto__in=produto_ids)
    }

    unidades_por_sigla: dict[str, UnidadeMedida] = {}
    for unidade in UnidadeMedida.objects.all():
        token = _normalize_unit_token(unidade.sigla)
        if token:
            unidades_por_sigla[token] = unidade

    formas_origem_validas = {
        (str(tipo).strip().upper(), int(id_origem))
        for tipo, id_origem in FormaPagamentoOrigem.objects.filter(ativo=True).values_list(
            "tipo_documento", "id_forma_origem"
        )
    }
    mapeamentos_ativos = {
        (str(item.tipo_documento).strip().upper(), int(item.id_forma_origem)): item.forma_pagamento
        for item in FormaPagamentoMapeamento.objects.select_related("forma_pagamento").filter(ativo=True)
    }

    clientes_padrao = list(Cliente.objects.filter(cliente_padrao=True).order_by("id_cliente")[:2])
    if len(clientes_padrao) > 1:
        raise ValueError("Consolidacao bloqueada: existe mais de um cliente padrao configurado.")
    cliente_padrao = clientes_padrao[0] if clientes_padrao else None

    erros_precheck: list[str] = []
    preparadas: list[dict[str, Any]] = []

    for stg_venda in candidatas:
        prefixo = f"{stg_venda.tipo_documento}#{stg_venda.id_legado}"
        erros_venda: list[str] = []

        motivos_bloqueantes_snapshot = sorted(_motivos_bloqueantes_consolidacao(stg_venda))
        if motivos_bloqueantes_snapshot and not _permite_override_motivos_reconciliacao(
            set(motivos_bloqueantes_snapshot),
            forcar_divergencia_formato,
        ):
            erros_venda.append(
                f"{prefixo}: divergencias_bloqueantes_reconciliacao "
                f"({', '.join(motivos_bloqueantes_snapshot)})"
            )

        itens = itens_por_venda.get(stg_venda.pk, [])
        pagamentos = pagamentos_por_venda.get(stg_venda.pk, [])
        if not itens or not pagamentos:
            erros_venda.append(f"{prefixo}: venda_sem_itens_ou_pagamentos")

        usuario = None
        if stg_venda.id_usuario_legado is None:
            erros_venda.append(f"{prefixo}: usuario_legado_ausente")
        else:
            usuario = usuarios_por_id.get(int(stg_venda.id_usuario_legado))
            if usuario is None:
                erros_venda.append(
                    f"{prefixo}: usuario_legado_nao_encontrado (id={stg_venda.id_usuario_legado})"
                )
            else:
                nome_stg = _normalize_identity_text(stg_venda.nome_usuario_legado)
                nome_sot = _normalize_identity_text(usuario.nome)
                if nome_stg and nome_sot and nome_stg != nome_sot:
                    erros_venda.append(
                        f"{prefixo}: usuario_nome_divergente (stg='{stg_venda.nome_usuario_legado}', sot='{usuario.nome}')"
                    )

        cliente = None
        if stg_venda.id_cliente_legado in (None, 0):
            if cliente_padrao is None:
                erros_venda.append(
                    f"{prefixo}: cliente_legado_zero_sem_cliente_padrao_configurado"
                )
            else:
                cliente = cliente_padrao
        else:
            cliente = clientes_por_id.get(int(stg_venda.id_cliente_legado))
            if cliente is None:
                erros_venda.append(
                    f"{prefixo}: cliente_nao_encontrado (id={stg_venda.id_cliente_legado})"
                )
            else:
                nome_stg = _normalize_identity_text(stg_venda.nome_cliente_legado)
                nome_sot = _normalize_identity_text(cliente.nome_cliente)
                nomes_compativeis, similaridade_cliente = _nomes_clientes_semelhantes(
                    stg_venda.nome_cliente_legado,
                    cliente.nome_cliente,
                )
                if nome_stg and nome_sot and not nomes_compativeis:
                    erros_venda.append(
                        f"{prefixo}: cliente_nome_divergente (similaridade={similaridade_cliente * 100:.1f}%, stg='{stg_venda.nome_cliente_legado}', sot='{cliente.nome_cliente}')"
                    )

        itens_convertidos: list[tuple[STG_ItemVenda, Produto, UnidadeMedida]] = []
        for item in itens:
            if item.id_produto_legado is None:
                erros_venda.append(f"{prefixo}: item_sem_id_produto")
                continue

            produto = produtos_por_id.get(int(item.id_produto_legado))
            if produto is None:
                erros_venda.append(
                    f"{prefixo}: produto_nao_encontrado (id={item.id_produto_legado})"
                )
                continue

            nomes_compativeis, similaridade_nome = _nomes_produtos_semelhantes(
                item.nome_produto_legado,
                produto.produto,
            )
            if not nomes_compativeis:
                erros_venda.append(
                    f"{prefixo}: produto_nome_divergente (id={item.id_produto_legado}, similaridade={similaridade_nome * 100:.1f}%, stg='{item.nome_produto_legado}', sot='{produto.produto}')"
                )
                continue

            unidade = _resolver_unidade_legado(
                unidade_legado=item.unidade_comercial_legado,
                unidades_por_sigla=unidades_por_sigla,
            )
            if unidade is None:
                erros_venda.append(
                    f"{prefixo}: unidade_legado_sem_mapeamento (valor='{item.unidade_comercial_legado}')"
                )
                continue

            itens_convertidos.append((item, produto, unidade))

        pagamentos_convertidos: list[tuple[FormaPagamento, STG_PagamentoVenda]] = []
        for pg in pagamentos:
            if pg.id_tipo_pagamento_legado is None:
                erros_venda.append(f"{prefixo}: pagamento_sem_id_forma_origem")
                continue

            chave = (stg_venda.tipo_documento, int(pg.id_tipo_pagamento_legado))
            if chave not in formas_origem_validas:
                erros_venda.append(
                    f"{prefixo}: forma_origem_nao_cadastrada (tipo={stg_venda.tipo_documento}, id_origem={pg.id_tipo_pagamento_legado})"
                )
                continue

            forma = mapeamentos_ativos.get(chave)
            if forma is None:
                erros_venda.append(
                    f"{prefixo}: mapeamento_forma_ausente (tipo={stg_venda.tipo_documento}, id_origem={pg.id_tipo_pagamento_legado})"
                )
                continue

            pagamentos_convertidos.append((forma, pg))

        if erros_venda:
            erros_precheck.extend(erros_venda)
            continue

        preparadas.append(
            {
                "stg_venda": stg_venda,
                "cliente": cliente,
                "usuario": usuario,
                "itens": itens_convertidos,
                "pagamentos": pagamentos_convertidos,
            }
        )

    if erros_precheck:
        bloqueios_precheck = _agrupar_erros_precheck_por_venda(erros_precheck)
        codigos_precheck: set[str] = set()
        for item in bloqueios_precheck:
            codigos_precheck.update(item.get("codigos") or [])

        if not _permite_override_precheck(codigos_precheck, forcar_divergencia_formato):
            raise ReconciliacaoBloqueioError(
                _formatar_erros_precheck(erros_precheck),
                codigo="consolidacao_bloqueada_precheck",
                bloqueios=bloqueios_precheck,
                permite_override=_codigos_precheck_apenas_formato(codigos_precheck),
            )

    inseridas = 0
    momento_consolidacao = timezone.now()
    with transaction.atomic():
        for item_venda in preparadas:
            stg_venda = item_venda["stg_venda"]
            venda = Venda.objects.create(
                id_legado=stg_venda.id_legado,
                tipo_documento=stg_venda.tipo_documento,
                data_venda=stg_venda.data_venda,
                hora_venda=stg_venda.hora_venda,
                status=stg_venda.status_venda or "",
                cliente=item_venda["cliente"],
                usuario=item_venda["usuario"],
                valor_total_documento=stg_venda.valor_final or Decimal("0"),
                momento_consolidacao=momento_consolidacao,
            )

            ItemVenda.objects.bulk_create(
                [
                    ItemVenda(
                        venda=venda,
                        produto=produto,
                        unidade_medida=unidade,
                        quantidade=stg_item.quantidade or Decimal("0"),
                        valor_unitario=stg_item.valor_unitario or Decimal("0"),
                        valor_total_item=stg_item.valor_total_calculado or Decimal("0"),
                        cancelado=stg_item.cancelado,
                    )
                    for stg_item, produto, unidade in item_venda["itens"]
                ],
                batch_size=2000,
            )

            PagamentoVenda.objects.bulk_create(
                [
                    PagamentoVenda(
                        venda=venda,
                        forma_pagamento=forma,
                        valor_pago=pg.valor_pago or Decimal("0"),
                        estorno=bool(pg.estorno),
                    )
                    for forma, pg in item_venda["pagamentos"]
                ],
                batch_size=2000,
            )

            inseridas += 1

        stg_vendas_removidas = STG_Venda.objects.count()
        stg_auditoria_removidas = STG_AuditoriaPlanilha.objects.count()
        STG_Venda.objects.all().delete()
        STG_AuditoriaPlanilha.objects.all().delete()

    return {
        "vendas_inseridas": inseridas,
        "vendas_ignoradas_duplicadas": ignoradas_duplicadas,
        "vendas_ignoradas_incompletas": 0,
        "detalhes_ignorados": [],
        "momento_consolidacao": momento_consolidacao.isoformat(),
        "stg_vendas_removidas": stg_vendas_removidas,
        "stg_auditoria_removidas": stg_auditoria_removidas,
    }
